#!/usr/bin/env python3
import sys
import shutil
import json
from pathlib import Path
from itertools import combinations
import openpyxl
import random

def generate_league(roosters_folder, num_divisions=None):
    """Generate league pairings and create match templates."""
    roosters_path = Path(roosters_folder)
    fixtures_data = {}
    
    # If path is a directory, look for Roosters subfolder
    if roosters_path.is_dir() and not any(roosters_path.glob("*.pdf")):
        potential_roosters = roosters_path / "Roosters"
        if potential_roosters.exists():
            roosters_path = potential_roosters
    
    if not roosters_path.exists():
        print(f"Error: Folder '{roosters_folder}' does not exist.")
        return
    
    # Check for division subfolders
    division_folders = [d for d in roosters_path.iterdir() if d.is_dir()]
    
    if division_folders:
        # Process each division separately
        for division_folder in division_folders:
            division_name = division_folder.name
            teams = [f.stem for f in division_folder.glob("*.pdf")]
            if len(teams) >= 2:
                print(f"\nProcessing division: {division_name}")
                division_fixtures = generate_division_league(teams, roosters_path.parent, division_name)
                if division_fixtures:
                    fixtures_data[division_name] = division_fixtures
        return
    
    # No divisions, process as single league or split into divisions
    teams = [f.stem for f in roosters_path.glob("*.pdf")]
    
    if len(teams) < 2:
        print("Need at least 2 teams to generate a league.")
        return
    
    # If num_divisions specified, split teams randomly
    if num_divisions and num_divisions > 1:
        if len(teams) % num_divisions != 0:
            print(f"Error: Cannot split {len(teams)} teams evenly into {num_divisions} divisions.")
            return
        
        # Shuffle and split teams
        random.shuffle(teams)
        teams_per_division = len(teams) // num_divisions
        
        for i in range(num_divisions):
            division_name = f"Division {i + 1}"
            division_teams = teams[i * teams_per_division:(i + 1) * teams_per_division]
            print(f"\nProcessing division: {division_name}")
            division_fixtures = generate_division_league(division_teams, roosters_path.parent, division_name)
            if division_fixtures:
                fixtures_data[division_name] = division_fixtures
    else:
        fixtures_data = generate_division_league(teams, roosters_path.parent, None)
    
    # Save fixtures to JSON
    fixtures_file = roosters_path.parent / "Fixtures" / "fixtures.json"
    with open(fixtures_file, 'w', encoding='utf-8') as f:
        json.dump(fixtures_data, f, indent=2, ensure_ascii=False)
    print(f"\nFixtures saved to: {fixtures_file}")

def generate_division_league(teams, parent_dir, division_name=None):
    """Generate league schedule for a division."""
    # Generate round-robin schedule (each team plays once per date)
    n = len(teams)
    if n % 2 == 1:
        teams.append("BYE")  # Add bye for odd number of teams
        n += 1
    
    schedule = []
    fixtures_json = {}
    
    for round_num in range(n - 1):
        round_matches = []
        for i in range(n // 2):
            team1 = teams[i]
            team2 = teams[n - 1 - i]
            if team1 != "BYE" and team2 != "BYE":
                round_matches.append((team1, team2))
        schedule.append(round_matches)
        # Rotate teams (keep first team fixed)
        teams = [teams[0]] + [teams[-1]] + teams[1:-1]
    
    total_dates = len(schedule)
    
    template_path = Path("samples/clean/Hoja Limpia Acta.xlsx")
    
    if not template_path.exists():
        print(f"Template file not found: {template_path}")
        return
    
    # Create date folders under Fixtures folder
    fixtures_dir = parent_dir / "Fixtures"
    fixtures_dir.mkdir(exist_ok=True)
    
    if division_name:
        base_dir = fixtures_dir / division_name
        base_dir.mkdir(exist_ok=True)
    else:
        base_dir = fixtures_dir
    
    # Wipe previous date folders
    for existing_folder in base_dir.glob("J*"):
        if existing_folder.is_dir():
            shutil.rmtree(existing_folder)
            print(f"Removed: {existing_folder}")
    
    for date_num, round_matches in enumerate(schedule, 1):
        date_folder = base_dir / f"J{date_num}"
        date_folder.mkdir(exist_ok=True)
        
        date_key = f"J{date_num}"
        fixtures_json[date_key] = []
        
        for i, (team1, team2) in enumerate(round_matches, 1):
            match_file = date_folder / f"Match_{i}_{team1}_vs_{team2}.xlsx"
            shutil.copy2(template_path, match_file)
            
            # Update team names in Excel file
            wb = openpyxl.load_workbook(match_file)
            ws = wb.active
            ws.cell(row=2, column=2, value=team1)
            ws.cell(row=2, column=3, value=team2)
            wb.save(match_file)
            
            # Add to fixtures JSON
            fixtures_json[date_key].append({
                "local": team1,
                "visitante": team2
            })
            
            print(f"Created: {match_file}")
    
    total_matches = sum(len(round_matches) for round_matches in schedule)
    print(f"\nLeague generated with {len([t for t in teams if t != 'BYE'])} teams, {total_matches} matches across {total_dates} dates.")
    
    return fixtures_json

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_league.py <roosters_folder_path> [--divisions N]")
        sys.exit(1)
    
    roosters_folder = sys.argv[1]
    num_divisions = None
    
    # Parse --divisions parameter
    if "--divisions" in sys.argv:
        idx = sys.argv.index("--divisions")
        if idx + 1 < len(sys.argv):
            try:
                num_divisions = int(sys.argv[idx + 1])
                if num_divisions <= 0:
                    print("Error: divisions must be a positive integer.")
                    sys.exit(1)
            except ValueError:
                print("Error: divisions must be a positive integer.")
                sys.exit(1)
    
    generate_league(roosters_folder, num_divisions)